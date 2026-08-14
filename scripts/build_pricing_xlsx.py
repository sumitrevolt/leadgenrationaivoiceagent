"""Build Niche_Pricing_Research.xlsx — top 25 niches + competitor pricing + benchmarks."""

import sys

sys.path.insert(0, r"C:\Users\Ratanshila\Documents\leadgenrationaivoiceagent")
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from app.marketing.voice_packages import VOICE_TIERS, lead_topup_price
from app.niches import NICHES

F = "Arial"
HDR_FILL = PatternFill("solid", start_color="1F4E79")
TIER_FILL = {"S": "C6EFCE", "A": "FFEB9C", "B": "DDEBF7"}
THIN = Border(*[Side(style="thin", color="B0B0B0")] * 4)

wb = Workbook()

# ---------------- Sheet 1: Top 25 Niches ----------------
ws = wb.active
ws.title = "Top 25 Niches"
ws["A1"] = "LeadGen AI — Top 25 Niches (research-finalized, June 2026)"
ws["A1"].font = Font(name=F, bold=True, size=14)
ws["A2"] = (
    "Two-tier model: hum B2B clients lete hain; unka agent unke end-customers (B2C ya B2B) ko call karta hai. Prices in INR."
)
ws["A2"].font = Font(name=F, italic=True, size=10)

headers = [
    "Rank",
    "Niche",
    "Tier",
    "End-Customer Type",
    "B2B Client (hamara customer)",
    "End Customer (agent kise call karega)",
    "Avg Ticket (INR)",
    "Lead Band (A/B/C)",
    "Voice Starter ₹/mo (10 leads)",
    "Voice Growth ₹/mo (30 leads)",
    "Voice Pro ₹/mo (60 leads)",
    "Top-up ₹ (10 extra leads)",
]
HR = 4
for c, h in enumerate(headers, 1):
    cell = ws.cell(row=HR, column=c, value=h)
    cell.font = Font(name=F, bold=True, color="FFFFFF", size=10)
    cell.fill = HDR_FILL
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    cell.border = THIN

tier_order = {"S": 0, "A": 1, "B": 2}
rows = sorted(NICHES.items(), key=lambda kv: (tier_order[kv[1]["tier"]],))
r = HR
for rank, (nid, cfg) in enumerate(rows, 1):
    r += 1
    band = str(cfg.get("lead_band") or "A").upper()
    _vt = {t["key"]: int(t["price_inr_month"].get(band, 0)) for t in VOICE_TIERS}
    vals = [
        rank,
        cfg["name"],
        cfg["tier"],
        cfg["target_type"].upper(),
        cfg["b2b_client"],
        cfg["end_customer"],
        cfg["avg_ticket_inr"],
        band,
        _vt.get("voice_starter", 0),
        _vt.get("voice_growth", 0),
        _vt.get("voice_pro", 0),
        lead_topup_price(band),
    ]
    for c, v in enumerate(vals, 1):
        cell = ws.cell(row=r, column=c, value=v)
        cell.font = Font(name=F, size=10)
        cell.border = THIN
        if c in (9, 10, 11, 12):
            cell.number_format = "₹#,##0"
        if c == 3:
            cell.fill = PatternFill("solid", start_color=TIER_FILL[cfg["tier"]])
            cell.alignment = Alignment(horizontal="center")
        if c in (1, 4):
            cell.alignment = Alignment(horizontal="center")
        if c in (5, 6, 7):
            cell.alignment = Alignment(wrap_text=True, vertical="top")
widths = [6, 30, 6, 12, 34, 44, 26, 13, 13, 12, 13, 13, 14]
for i, w in enumerate(widths, 1):
    ws.column_dimensions[get_column_letter(i)].width = w
ws.freeze_panes = "A5"

# ---------------- Sheet 2: Competitor Pricing ----------------
ws2 = wb.create_sheet("Competitor Pricing")
ws2["A1"] = "AI Voice Agent Platforms — Pricing Scan (June 2026)"
ws2["A1"].font = Font(name=F, bold=True, size=14)
comp_headers = [
    "Platform",
    "Type",
    "Pricing Model",
    "Effective Price",
    "India-Ready?",
    "Notes / Source",
]
comp = [
    [
        "Vapi",
        "Global platform",
        "Usage: $0.05/min platform + model costs at-cost",
        "$0.07–0.33/min all-in",
        "No (TRAI nahi; SIP workaround)",
        "vapi.ai/pricing; cekura.ai",
    ],
    [
        "Retell AI",
        "Global platform",
        "Pay-as-you-go bundle",
        "$0.07–0.31/min; ~$300/mo per 3k min",
        "No (BYO SIP)",
        "retellai.com/pricing",
    ],
    [
        "Bland AI",
        "Global platform",
        "Tiered subscription, bundled",
        "$0.11–0.14/min",
        "No",
        "ringg.ai; cloudtalk.io",
    ],
    [
        "Synthflow",
        "Global platform",
        "PAYG + white-label $2,000/mo",
        "$0.11–0.24/min",
        "No",
        "zeeg.me; famulor.io",
    ],
    [
        "Air.ai",
        "Enterprise",
        "$25K–100K license + per-min",
        "$0.11–0.32/min",
        "No",
        "lindy.ai/blog",
    ],
    [
        "Vogent",
        "Global platform",
        "Flat per-minute",
        "$0.09/min (vol. $0.04–0.05)",
        "No",
        "blog.vogent.ai",
    ],
    [
        "ElevenLabs Agents",
        "Global platform",
        "Tiered per-minute + LLM extra",
        "$0.08–0.12/min +10-30%",
        "Partial (SIP)",
        "elevenlabs.io/pricing/agents",
    ],
    [
        "Dograh",
        "Open-source",
        "Self-hosted, BYOK — $0 platform",
        "<$2,000/mo @20k min (vs $6k Vapi)",
        "YES (Exotel/Plivo SIP)",
        "github.com/dograh-hq; blog.dograh.com",
    ],
    [
        "Exotel AgentStream",
        "India enterprise",
        "Custom quote; telephony ₹0.8–1/min",
        "Quote-based",
        "YES (DLT-native)",
        "exotel.com",
    ],
    [
        "CallHippo AI",
        "India SMB",
        "Monthly bundles",
        "$199/2,500min → $799/7,500min; $0.16/min over",
        "YES",
        "emitrr.com; g2.com",
    ],
    [
        "Vodex",
        "India outbound",
        "Monthly bundles, connected-calls billed",
        "$100–500/mo (~₹11/min eff.)",
        "YES",
        "vodex.ai/pricing",
    ],
    [
        "SquadStack",
        "India service",
        "Pay-per-qualified-lead (AI+human)",
        "Per-lead quote; 3-attempt/connect billing",
        "YES",
        "squadstack.ai",
    ],
    ["Bolna.ai", "India platform", "Per-minute", "~₹5.52/min", "YES", "caller.digital roundup"],
    [
        "Sarvam Samvaad",
        "India models",
        "API: STT ₹30/hr, TTS ₹15–30/10k chars",
        "~₹1/min model cost claim",
        "YES",
        "sarvam.ai/api-pricing",
    ],
    [
        "smallest.ai Atoms",
        "India platform",
        "Plans from $5/mo + usage",
        "~$0.01/min infra + TTS $0.028/min",
        "YES",
        "smallest.ai/pricing",
    ],
    [
        "Knowlarity",
        "India telephony",
        "Per agent per month",
        "₹1,999–3,499/agent/mo",
        "YES",
        "knowlarity.com/pricing",
    ],
    [
        "MyOperator",
        "India telephony",
        "Monthly + onboarding",
        "₹2,500/mo + ₹20k setup",
        "YES",
        "myoperator.com",
    ],
    [
        "DigiBrood",
        "India bulk AI",
        "Per call",
        "₹0.50/call se (bulk 22–28p)",
        "YES",
        "digibrood.in",
    ],
]
for c, h in enumerate(comp_headers, 1):
    cell = ws2.cell(row=3, column=c, value=h)
    cell.font = Font(name=F, bold=True, color="FFFFFF", size=10)
    cell.fill = HDR_FILL
    cell.border = THIN
for i, row in enumerate(comp, 4):
    for c, v in enumerate(row, 1):
        cell = ws2.cell(row=i, column=c, value=v)
        cell.font = Font(name=F, size=10)
        cell.border = THIN
        cell.alignment = Alignment(wrap_text=True, vertical="top")
for i, w in enumerate([20, 16, 38, 34, 26, 34], 1):
    ws2.column_dimensions[get_column_letter(i)].width = w
ws2.freeze_panes = "A4"

# ---------------- Sheet 3: India Benchmarks ----------------
ws3 = wb.create_sheet("India CPL Benchmarks")
ws3["A1"] = "India Cost-Per-Lead / Qualified-Lead Benchmarks (2025-26)"
ws3["A1"].font = Font(name=F, bold=True, size=14)
b_headers = ["Industry", "Raw Lead CPL (INR)", "Qualified / Outcome Value (INR)", "Source"]
bench = [
    [
        "Real estate (resi, Meta/Google)",
        "₹150–1,500 (afford. ₹500–1,300; luxury ₹2,000–5,000)",
        "Site-visit-done ~₹7,500 (verified plausible)",
        "orynet.com; tatva.digital",
    ],
    [
        "Solar residential",
        "₹150–200 (Google case study)",
        "Call-center-gen lead ~₹2,100 ($25)",
        "tej9.com; solarreviews.com",
    ],
    [
        "Healthcare raw (clinic/dental)",
        "₹150–1,200 (avg ₹300–700)",
        "CPQL: diagnostics ₹900, aesthetics ₹1,800",
        "owlclaw.com; ichelon CPQL 2026",
    ],
    ["IVF", "(raw upar)", "CPQL ~₹2,200 (best ₹880–1,120)", "ichelonconsulting.com"],
    ["Hair transplant", "(raw upar)", "CPQL ~₹4,100 (best ₹2,200)", "ichelonconsulting.com"],
    [
        "Insurance",
        "₹500–1,500 (good ₹300–500)",
        "Commission 35–40% first-yr premium",
        "owlclaw.com; ebharat.com",
    ],
    ["Loans (DSA)", "Disbursal-based", "Payout 1–2.5% of disbursed (₹3L → ₹3–7.5K)", "ruloans.com"],
    [
        "Study abroad",
        "₹800–2,000",
        "₹2–3L commission/enrolled (CONFIRMED)",
        "orynet.com; studentsherald.com",
    ],
    ["Interior/modular kitchen", "₹100–400 (Meta local)", "—", "blusteak.com"],
    ["Car dealers", "₹435–560 per test-drive lead", "—", "blusteak.com case study"],
    ["Coaching NEET/JEE metro", "₹1,000–2,200/enquiry", "Fees ₹1.5–2.5L/yr", "mgiwebzone.com"],
    ["B2B IndiaMART BuyLead", "₹16–24 effective (shared)", "Export ~₹212", "refrens.com"],
    ["Legal/CA", "₹700–4,000", "—", "orynet.com"],
    ["B2B appointment (LinkedIn)", "—", "₹2,000–12,000/qualified appointment", "leadshuttle.com"],
    [
        "Human telecaller baseline",
        "₹55–70/contact (vendor est., plausible)",
        "Seat ₹8K–37.5K/mo; salary ₹17K/mo",
        "callcentersindia; jooble",
    ],
    [
        "AI calling baseline",
        "₹14–18/contact; ₹3–15/connected-min",
        "Case: ₹24.93/qualified lead",
        "caller.digital; getalchemystai",
    ],
]
for c, h in enumerate(b_headers, 1):
    cell = ws3.cell(row=3, column=c, value=h)
    cell.font = Font(name=F, bold=True, color="FFFFFF", size=10)
    cell.fill = HDR_FILL
    cell.border = THIN
for i, row in enumerate(bench, 4):
    for c, v in enumerate(row, 1):
        cell = ws3.cell(row=i, column=c, value=v)
        cell.font = Font(name=F, size=10)
        cell.border = THIN
        cell.alignment = Alignment(wrap_text=True, vertical="top")
for i, w in enumerate([34, 38, 42, 30], 1):
    ws3.column_dimensions[get_column_letter(i)].width = w
ws3.freeze_panes = "A4"

# ---------------- Sheet 4: Pricing Strategy ----------------
ws4 = wb.create_sheet("Pricing Strategy")
ws4["A1"] = "Pricing Strategy — kaise charge karein"
ws4["A1"].font = Font(name=F, bold=True, size=14)
strat = [
    ["", ""],
    [
        "PRINCIPLE",
        "Per-qualified-lead pricing (market gap: India me koi AI player 'leads delivered' package nahi bechta — sab per-minute bechte hain)",
    ],
    ["", ""],
    ["Band by ticket size", "Low-ticket (B2B suppliers, movers, used-cars): ₹150–800/QL"],
    ["", "Mid-ticket (solar resi, insurance, coaching, dental): ₹300–1,500/QL"],
    ["", "High-ticket (RE, study-abroad, hair/IVF, luxury): ₹800–6,000/QL"],
    ["", "Appointments/site-visits = 2–5x lead price (RE site visit market ~₹7,500)"],
    ["", ""],
    [
        "COST FLOOR (hamara)",
        "AI cost ~₹14–18/contact; ~8–15 contacts per qualified lead → cost ₹120–270/QL",
    ],
    [
        "Example margin @ ₹500 QL",
        '=ROUND((500-200)/500*100,0) & "% gross margin (cost ₹200 assumed)"',
    ],
    ["", ""],
    [
        "MONTHLY PLANS (suggested)",
        "Starter ₹8K–25K/mo (niche-wise, capped QLs) — agency retainers ₹15–50K se neeche position karo",
    ],
    ["", "Growth: starter × 2 (2x lead cap + WhatsApp/CRM integrations)"],
    [
        "",
        "Scale/white-label: custom — digital agencies ko reseller banao (Synthflow white-label $2,000/mo hai — hum sasta de sakte)",
    ],
    ["", ""],
    [
        "COMPLIANCE (zaroori)",
        "140-series number + DLT registration + DND scrubbing + 10am–7pm calling + AI-call disclosure (TCCCPR 2026) — penalty ₹10L tak (CONFIRMED)",
    ],
    ["", ""],
    [
        "TIER ROLLOUT",
        "S-tier (8 niches) pehle becho — RE, study-abroad, home-loans, solar, insurance, coaching. A-tier expand. B-tier volume/strategic.",
    ],
]
r0 = 2
for i, (a, b) in enumerate(strat, r0):
    ca = ws4.cell(row=i, column=1, value=a)
    ca.font = Font(
        name=F,
        bold=bool(a and a.isupper() or a.startswith("Band") or a.startswith("Example")),
        size=10,
    )
    if a == "Example margin @ ₹500 QL":
        cb = ws4.cell(
            row=i,
            column=2,
            value='=ROUND((500-200)/500*100,0)&"% gross margin (cost ₹200 assumed)"',
        )
    else:
        cb = ws4.cell(row=i, column=2, value=b)
    cb.font = Font(name=F, size=10)
    cb.alignment = Alignment(wrap_text=True, vertical="top")
ws4.column_dimensions["A"].width = 28
ws4.column_dimensions["B"].width = 110

out = r"C:\Users\Ratanshila\Documents\leadgenrationaivoiceagent\Niche_Pricing_Research.xlsx"
wb.save(out)
print("saved", out)
